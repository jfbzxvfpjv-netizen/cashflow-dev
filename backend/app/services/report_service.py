"""
Servicio de Informes — M8
Generación de informes de cierre de sesión y período libre
en formato PDF (WeasyPrint) y Excel (openpyxl).
"""
import io
from datetime import date, datetime
from decimal import Decimal
from typing import Optional, List, Dict, Any

from sqlalchemy import cast, Date, func
from sqlalchemy.orm import Session

from app.models.cash_flow import (
    CashSession, Transaction, SystemConfig, TransactionSignature, TransactionProject
)
from app.models.catalogs import (
    TransactionCategory, TransactionSubcategory,
    Supplier, Employee, Partner, Project, Work, Vehicle
)
from app.models.user import User


class ReportService:
    """Genera informes en PDF y Excel."""

    @staticmethod
    def _get_contraparte(db: Session, t: Transaction) -> str:
        """Obtiene el nombre de la contraparte de una transacción."""
        if t.counterparty_free:
            return t.counterparty_free
        if t.supplier_id:
            s = db.query(Supplier).filter(Supplier.id == t.supplier_id).first()
            return s.name if s else ''
        if t.employee_id:
            e = db.query(Employee).filter(Employee.id == t.employee_id).first()
            return e.full_name if e else ''
        if t.partner_id:
            p = db.query(Partner).filter(Partner.id == t.partner_id).first()
            return p.full_name if p else ''
        return ''

    @staticmethod
    def _format_amount(amount) -> str:
        """Formatea un importe numérico con separadores de miles."""
        try:
            val = float(amount)
            if val == int(val):
                return f"{int(val):,}".replace(",", ".")
            return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except (ValueError, TypeError):
            return "0"

    @classmethod
    def get_session_data(cls, db: Session, session_id: int) -> Dict[str, Any]:
        """Obtiene todos los datos necesarios para el informe de sesión."""
        session = db.query(CashSession).filter(CashSession.id == session_id).first()
        if not session:
            return None

        user = db.query(User).filter(User.id == session.user_id).first()

        transactions = db.query(Transaction).filter(
            Transaction.session_id == session_id
        ).order_by(Transaction.created_at).all()

        items = []
        total_ingresos = Decimal('0')
        total_egresos = Decimal('0')

        for t in transactions:
            cat = db.query(TransactionCategory).filter(
                TransactionCategory.id == t.category_id
            ).first()

            contraparte = cls._get_contraparte(db, t)

            estado = 'Aprobada'
            if t.cancelled:
                estado = 'Anulada'
            elif t.approval_status == 'pending_approval':
                estado = 'Pendiente'
            elif t.approval_status == 'authorized':
                estado = 'Autorizada'
            elif t.approval_status == 'rejected':
                estado = 'Rechazada'

            if not t.cancelled and t.approval_status == 'approved':
                if t.type == 'income':
                    total_ingresos += t.amount
                else:
                    total_egresos += t.amount

            items.append({
                "fecha": t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else '',
                "referencia": t.reference_number or '',
                "concepto": t.concept or '',
                "categoria": cat.name if cat else '',
                "contraparte": contraparte,
                "tipo": t.type,
                "importe": float(t.amount),
                "estado": estado,
                "cancelled": t.cancelled
            })

        return {
            "session": {
                "id": session.id,
                "delegacion": session.delegacion,
                "gestor": user.full_name if user else '',
                "opened_at": session.opened_at.strftime("%d/%m/%Y %H:%M") if session.opened_at else '',
                "closed_at": session.closed_at.strftime("%d/%m/%Y %H:%M") if session.closed_at else 'Abierta',
                "opening_balance": float(session.opening_balance),
                "closing_balance": float(session.closing_balance) if session.closing_balance is not None else None,
                "status": session.status
            },
            "transactions": items,
            "total_ingresos": float(total_ingresos),
            "total_egresos": float(total_egresos),
            "diferencia": float(total_ingresos - total_egresos)
        }

    @classmethod
    def get_subcategory_data(
        cls, db: Session,
        delegacion: Optional[str],
        date_start: date,
        date_end: date,
        subcategory_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """Datos del informe por subcategoria.
        Sin subcategory_id -> resumen agrupado por subcategoria.
        Con subcategory_id -> detalle de movimientos de esa subcategoria."""
        from app.models import TransactionSubcategory
        query = db.query(Transaction).filter(
            cast(Transaction.created_at, Date) >= date_start,
            cast(Transaction.created_at, Date) <= date_end
        )
        if delegacion and delegacion != 'Consolidado':
            query = query.filter(Transaction.delegacion == delegacion)
        if subcategory_id:
            query = query.filter(Transaction.subcategory_id == subcategory_id)
        transactions = query.order_by(Transaction.created_at).all()
        periodo = {
            "date_start": date_start.strftime("%d/%m/%Y"),
            "date_end": date_end.strftime("%d/%m/%Y"),
            "delegacion": delegacion or "Todas",
        }
        if subcategory_id:
            sub = db.query(TransactionSubcategory).filter(
                TransactionSubcategory.id == subcategory_id).first()
            cat = None
            if sub:
                cat = db.query(TransactionCategory).filter(
                    TransactionCategory.id == sub.category_id).first()
            items = []
            total_ingresos = Decimal('0')
            total_egresos = Decimal('0')
            for t in transactions:
                contraparte = cls._get_contraparte(db, t)
                estado = 'Aprobada'
                if t.cancelled:
                    estado = 'Anulada'
                elif t.approval_status == 'pending_approval':
                    estado = 'Pendiente'
                elif t.approval_status == 'authorized':
                    estado = 'Autorizada'
                elif t.approval_status == 'rejected':
                    estado = 'Rechazada'
                if not t.cancelled and t.approval_status == 'approved':
                    if t.type == 'income':
                        total_ingresos += t.amount
                    else:
                        total_egresos += t.amount
                items.append({
                    "fecha": t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else '',
                    "referencia": t.reference_number or '',
                    "concepto": t.concept or '',
                    "contraparte": contraparte,
                    "tipo": t.type,
                    "importe": float(t.amount),
                    "estado": estado,
                    "delegacion": t.delegacion,
                    "cancelled": t.cancelled,
                })
            return {
                "modo": "detalle",
                "periodo": periodo,
                "subcategoria": sub.name if sub else str(subcategory_id),
                "categoria": cat.name if cat else '',
                "transactions": items,
                "total_ingresos": float(total_ingresos),
                "total_egresos": float(total_egresos),
                "saldo_neto": float(total_ingresos - total_egresos),
            }
        grupos = {}
        for t in transactions:
            if t.cancelled or t.approval_status != 'approved':
                continue
            sid = t.subcategory_id
            if sid not in grupos:
                sub = db.query(TransactionSubcategory).filter(
                    TransactionSubcategory.id == sid).first() if sid else None
                cat = None
                if sub:
                    cat = db.query(TransactionCategory).filter(
                        TransactionCategory.id == sub.category_id).first()
                grupos[sid] = {
                    "subcategoria": sub.name if sub else 'Sin subcategoria',
                    "categoria": cat.name if cat else '',
                    "num": 0,
                    "ingresos": Decimal('0'),
                    "egresos": Decimal('0'),
                }
            g = grupos[sid]
            g["num"] += 1
            if t.type == 'income':
                g["ingresos"] += t.amount
            else:
                g["egresos"] += t.amount
        rows = []
        for g in grupos.values():
            rows.append({
                "subcategoria": g["subcategoria"],
                "categoria": g["categoria"],
                "num": g["num"],
                "ingresos": float(g["ingresos"]),
                "egresos": float(g["egresos"]),
                "neto": float(g["ingresos"] - g["egresos"]),
            })
        rows.sort(key=lambda r: r["egresos"], reverse=True)
        return {
            "modo": "resumen",
            "periodo": periodo,
            "rows": rows,
            "total_ingresos": float(sum(Decimal(str(r["ingresos"])) for r in rows)),
            "total_egresos": float(sum(Decimal(str(r["egresos"])) for r in rows)),
        }

    @classmethod
    def generate_subcategory_pdf(
        cls, db: Session,
        delegacion: Optional[str],
        date_start: date,
        date_end: date,
        subcategory_id: Optional[int] = None
    ) -> bytes:
        """Informe por subcategoria en PDF (resumen o detalle)."""
        data = cls.get_subcategory_data(db, delegacion, date_start, date_end, subcategory_id)
        p = data["periodo"]
        if data["modo"] == "detalle":
            titulo = "Detalle por Subcategoria"
            subtitulo = f"{data['categoria']} / {data['subcategoria']}"
            rows_html = ""
            for t in data["transactions"]:
                color = "#16a34a" if t["tipo"] == "income" else "#dc2626"
                if t["cancelled"]:
                    color = "#9ca3af"
                signo = "+" if t["tipo"] == "income" else "-"
                rows_html += f"""
                <tr style="{'text-decoration:line-through;color:#9ca3af;' if t['cancelled'] else ''}">
                    <td>{t['fecha']}</td>
                    <td>{t['referencia']}</td>
                    <td>{t['concepto'][:60]}</td>
                    <td>{t['contraparte'][:30]}</td>
                    <td style="color:{color};text-align:right;font-weight:600;">{signo} {cls._format_amount(t['importe'])} XAF</td>
                    <td>{t['delegacion']}</td>
                    <td>{t['estado']}</td>
                </tr>"""
            tabla = f"""
            <table>
                <thead><tr>
                    <th>Fecha/Hora</th><th>Ref.</th><th>Concepto</th><th>Contraparte</th>
                    <th style="text-align:right">Importe</th><th>Deleg.</th><th>Estado</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
            <div class="totals">
                <div style="color:#16a34a;">Total ingresos: <strong>{cls._format_amount(data['total_ingresos'])} XAF</strong></div>
                <div style="color:#dc2626;">Total egresos: <strong>{cls._format_amount(data['total_egresos'])} XAF</strong></div>
                <div style="color:#1e40af;font-size:13px;">Saldo neto: <strong>{cls._format_amount(data['saldo_neto'])} XAF</strong></div>
            </div>"""
        else:
            titulo = "Resumen por Subcategoria"
            subtitulo = f"{len(data['rows'])} subcategorias con movimiento"
            rows_html = ""
            for r in data["rows"]:
                rows_html += f"""
                <tr>
                    <td>{r['subcategoria']}</td>
                    <td>{r['categoria']}</td>
                    <td style="text-align:center">{r['num']}</td>
                    <td style="text-align:right;color:#16a34a;">{cls._format_amount(r['ingresos'])}</td>
                    <td style="text-align:right;color:#dc2626;">{cls._format_amount(r['egresos'])}</td>
                    <td style="text-align:right;font-weight:600;">{cls._format_amount(r['neto'])}</td>
                </tr>"""
            tabla = f"""
            <table>
                <thead><tr>
                    <th>Subcategoria</th><th>Categoria</th><th style="text-align:center">N mov.</th>
                    <th style="text-align:right">Ingresos</th><th style="text-align:right">Egresos</th>
                    <th style="text-align:right">Neto</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
            <div class="totals">
                <div style="color:#16a34a;">Total ingresos: <strong>{cls._format_amount(data['total_ingresos'])} XAF</strong></div>
                <div style="color:#dc2626;">Total egresos: <strong>{cls._format_amount(data['total_egresos'])} XAF</strong></div>
            </div>"""
        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    @page {{ size: A4 landscape; margin: 15mm; }}
    body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10px; color: #1e293b; }}
    h1 {{ font-size: 16px; color: #1e40af; margin-bottom: 5px; }}
    h2 {{ font-size: 12px; color: #475569; margin-top: 0; }}
    .info {{ background: #f1f5f9; padding: 8px 12px; border-radius: 4px; margin-bottom: 15px; font-size: 10px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
    th {{ background: #1e40af; color: white; padding: 6px 8px; text-align: left; font-size: 9px; }}
    td {{ padding: 5px 8px; border-bottom: 1px solid #e2e8f0; font-size: 9px; }}
    tr:nth-child(even) {{ background: #f8fafc; }}
    .totals {{ margin-top: 15px; text-align: right; }}
    .totals div {{ margin: 3px 0; font-size: 11px; }}
    .footer {{ margin-top: 20px; font-size: 8px; color: #94a3b8; text-align: center; }}
</style></head><body>
    <h1>{titulo}</h1>
    <h2>{p['date_start']} &mdash; {p['date_end']} &nbsp;|&nbsp; {subtitulo}</h2>
    <div class="info">
        <strong>Delegacion:</strong> {p['delegacion']}
    </div>
    {tabla}
    <div class="footer">
        Generado automaticamente &mdash; {datetime.now().strftime('%d/%m/%Y %H:%M')} &mdash; Sistema de Gestion de Flujo de Caja
    </div>
</body></html>"""
        try:
            from weasyprint import HTML as WeasyHTML
            return WeasyHTML(string=html).write_pdf()
        except Exception:
            return html.encode('utf-8')

    @classmethod
    def generate_subcategory_xlsx(
        cls, db: Session,
        delegacion: Optional[str],
        date_start: date,
        date_end: date,
        subcategory_id: Optional[int] = None
    ) -> bytes:
        """Informe por subcategoria en Excel (resumen o detalle)."""
        data = cls.get_subcategory_data(db, delegacion, date_start, date_end, subcategory_id)
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        wb = Workbook()
        ws = wb.active
        header_font = Font(bold=True, size=14, color="1E40AF")
        th_font = Font(bold=True, size=9, color="FFFFFF")
        th_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        border = Border(bottom=Side(style='thin', color='E2E8F0'))
        p = data["periodo"]
        if data["modo"] == "detalle":
            ws.title = "Detalle subcategoria"
            ws['A1'] = f"Detalle por Subcategoria: {data['categoria']} / {data['subcategoria']}"
            ws['A1'].font = header_font
            ws['A3'] = "Periodo:"; ws['B3'] = f"{p['date_start']} - {p['date_end']}"
            ws['A4'] = "Delegacion:"; ws['B4'] = p['delegacion']
            for cell in ['A3', 'A4']:
                ws[cell].font = Font(bold=True, size=9)
            headers = ['Fecha/Hora', 'Referencia', 'Concepto', 'Contraparte', 'Importe', 'Delegacion', 'Estado']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=6, column=col, value=h)
                c.font = th_font; c.fill = th_fill; c.alignment = Alignment(horizontal='center')
            for i, t in enumerate(data["transactions"], 7):
                ws.cell(row=i, column=1, value=t['fecha'])
                ws.cell(row=i, column=2, value=t['referencia'])
                ws.cell(row=i, column=3, value=t['concepto'])
                ws.cell(row=i, column=4, value=t['contraparte'])
                amt = ws.cell(row=i, column=5, value=t['importe'])
                amt.number_format = '#,##0'
                amt.font = Font(color="16A34A" if t['tipo'] == 'income' else "DC2626", bold=True)
                ws.cell(row=i, column=6, value=t['delegacion'])
                ws.cell(row=i, column=7, value=t['estado'])
                for col in range(1, 8):
                    ws.cell(row=i, column=col).border = border
            last = 7 + len(data["transactions"])
            ws.cell(row=last+1, column=4, value="Total ingresos:").font = Font(bold=True)
            c = ws.cell(row=last+1, column=5, value=data['total_ingresos']); c.font = Font(color="16A34A", bold=True); c.number_format = '#,##0'
            ws.cell(row=last+2, column=4, value="Total egresos:").font = Font(bold=True)
            c = ws.cell(row=last+2, column=5, value=data['total_egresos']); c.font = Font(color="DC2626", bold=True); c.number_format = '#,##0'
            ws.cell(row=last+3, column=4, value="Saldo neto:").font = Font(bold=True, size=11)
            c = ws.cell(row=last+3, column=5, value=data['saldo_neto']); c.font = Font(bold=True, size=11, color="1E40AF"); c.number_format = '#,##0'
            widths = [18, 12, 40, 25, 16, 12, 12]
        else:
            ws.title = "Resumen subcategoria"
            ws['A1'] = "Resumen por Subcategoria"
            ws['A1'].font = header_font
            ws['A3'] = "Periodo:"; ws['B3'] = f"{p['date_start']} - {p['date_end']}"
            ws['A4'] = "Delegacion:"; ws['B4'] = p['delegacion']
            for cell in ['A3', 'A4']:
                ws[cell].font = Font(bold=True, size=9)
            headers = ['Subcategoria', 'Categoria', 'N mov.', 'Ingresos', 'Egresos', 'Neto']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=6, column=col, value=h)
                c.font = th_font; c.fill = th_fill; c.alignment = Alignment(horizontal='center')
            for i, r in enumerate(data["rows"], 7):
                ws.cell(row=i, column=1, value=r['subcategoria'])
                ws.cell(row=i, column=2, value=r['categoria'])
                ws.cell(row=i, column=3, value=r['num']).alignment = Alignment(horizontal='center')
                for col, key, col_color in [(4, 'ingresos', '16A34A'), (5, 'egresos', 'DC2626'), (6, 'neto', '1E40AF')]:
                    c = ws.cell(row=i, column=col, value=r[key])
                    c.number_format = '#,##0'
                    if col in (4, 5):
                        c.font = Font(color=col_color)
                    else:
                        c.font = Font(bold=True, color=col_color)
                for col in range(1, 7):
                    ws.cell(row=i, column=col).border = border
            last = 7 + len(data["rows"])
            ws.cell(row=last+1, column=2, value="TOTAL").font = Font(bold=True)
            c = ws.cell(row=last+1, column=4, value=data['total_ingresos']); c.font = Font(color="16A34A", bold=True); c.number_format = '#,##0'
            c = ws.cell(row=last+1, column=5, value=data['total_egresos']); c.font = Font(color="DC2626", bold=True); c.number_format = '#,##0'
            widths = [26, 22, 10, 16, 16, 16]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+col)].width = w
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def get_period_data(
        cls, db: Session,
        delegacion: Optional[str],
        date_start: date,
        date_end: date,
        category_id: Optional[int] = None,
        project_id: Optional[int] = None
    ) -> Dict[str, Any]:
        """Obtiene datos para el informe de período libre."""
        query = db.query(Transaction).filter(
            cast(Transaction.created_at, Date) >= date_start,
            cast(Transaction.created_at, Date) <= date_end
        )
        if delegacion and delegacion != 'Consolidado':
            query = query.filter(Transaction.delegacion == delegacion)
        if category_id:
            query = query.filter(Transaction.category_id == category_id)

        transactions = query.order_by(Transaction.created_at).all()

        items = []
        total_ingresos = Decimal('0')
        total_egresos = Decimal('0')

        for t in transactions:
            cat = db.query(TransactionCategory).filter(
                TransactionCategory.id == t.category_id
            ).first()
            contraparte = cls._get_contraparte(db, t)

            estado = 'Aprobada'
            if t.cancelled:
                estado = 'Anulada'
            elif t.approval_status == 'pending_approval':
                estado = 'Pendiente'
            elif t.approval_status == 'authorized':
                estado = 'Autorizada'
            elif t.approval_status == 'rejected':
                estado = 'Rechazada'

            if not t.cancelled and t.approval_status == 'approved':
                if t.type == 'income':
                    total_ingresos += t.amount
                else:
                    total_egresos += t.amount

            items.append({
                "fecha": t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else '',
                "referencia": t.reference_number or '',
                "concepto": t.concept or '',
                "categoria": cat.name if cat else '',
                "contraparte": contraparte,
                "tipo": t.type,
                "importe": float(t.amount),
                "estado": estado,
                "delegacion": t.delegacion,
                "cancelled": t.cancelled
            })

        filtros_texto = []
        if delegacion:
            filtros_texto.append(f"Delegación: {delegacion}")
        if category_id:
            cat = db.query(TransactionCategory).filter(TransactionCategory.id == category_id).first()
            filtros_texto.append(f"Categoría: {cat.name if cat else category_id}")

        return {
            "periodo": {
                "date_start": date_start.strftime("%d/%m/%Y"),
                "date_end": date_end.strftime("%d/%m/%Y"),
                "delegacion": delegacion or "Todas",
                "filtros": ", ".join(filtros_texto) if filtros_texto else "Sin filtros"
            },
            "transactions": items,
            "total_ingresos": float(total_ingresos),
            "total_egresos": float(total_egresos),
            "saldo_neto": float(total_ingresos - total_egresos)
        }

    @classmethod
    def generate_session_pdf(cls, db: Session, session_id: int) -> Optional[bytes]:
        """Genera informe de cierre de sesión en PDF."""
        data = cls.get_session_data(db, session_id)
        if not data:
            return None

        s = data["session"]
        rows_html = ""
        for t in data["transactions"]:
            color = "#16a34a" if t["tipo"] == "income" else "#dc2626"
            if t["cancelled"]:
                color = "#9ca3af"
            signo = "+" if t["tipo"] == "income" else "-"
            rows_html += f"""
            <tr style="{'text-decoration:line-through;color:#9ca3af;' if t['cancelled'] else ''}">
                <td>{t['fecha']}</td>
                <td>{t['referencia']}</td>
                <td>{t['concepto'][:60]}</td>
                <td>{t['categoria']}</td>
                <td>{t['contraparte'][:30]}</td>
                <td style="color:{color};text-align:right;font-weight:600;">{signo} {cls._format_amount(t['importe'])} XAF</td>
                <td>{t['estado']}</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    @page {{ size: A4 landscape; margin: 15mm; }}
    body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10px; color: #1e293b; }}
    h1 {{ font-size: 16px; color: #1e40af; margin-bottom: 5px; }}
    h2 {{ font-size: 12px; color: #475569; margin-top: 0; }}
    .header-grid {{ display: flex; justify-content: space-between; margin-bottom: 15px; }}
    .header-box {{ background: #f1f5f9; padding: 8px 12px; border-radius: 4px; }}
    .header-box span {{ display: block; font-size: 9px; color: #64748b; }}
    .header-box strong {{ font-size: 13px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
    th {{ background: #1e40af; color: white; padding: 6px 8px; text-align: left; font-size: 9px; }}
    td {{ padding: 5px 8px; border-bottom: 1px solid #e2e8f0; font-size: 9px; }}
    tr:nth-child(even) {{ background: #f8fafc; }}
    .totals {{ margin-top: 15px; text-align: right; }}
    .totals div {{ margin: 3px 0; font-size: 11px; }}
    .footer {{ margin-top: 20px; font-size: 8px; color: #94a3b8; text-align: center; }}
</style></head><body>
    <h1>Informe de Cierre de Sesión</h1>
    <h2>Delegación: {s['delegacion']} — Sesión #{s['id']}</h2>

    <div style="display:flex;gap:20px;margin-bottom:15px;">
        <div class="header-box"><span>Gestor</span><strong>{s['gestor']}</strong></div>
        <div class="header-box"><span>Apertura</span><strong>{s['opened_at']}</strong></div>
        <div class="header-box"><span>Cierre</span><strong>{s['closed_at']}</strong></div>
        <div class="header-box"><span>Saldo apertura</span><strong>{cls._format_amount(s['opening_balance'])} XAF</strong></div>
        <div class="header-box"><span>Saldo cierre</span><strong>{cls._format_amount(s['closing_balance']) if s['closing_balance'] is not None else 'N/A'} XAF</strong></div>
    </div>

    <table>
        <thead><tr>
            <th>Fecha/Hora</th><th>Ref.</th><th>Concepto</th>
            <th>Categoría</th><th>Contraparte</th><th style="text-align:right">Importe</th><th>Estado</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
    </table>

    <div class="totals">
        <div style="color:#16a34a;">Total ingresos: <strong>{cls._format_amount(data['total_ingresos'])} XAF</strong></div>
        <div style="color:#dc2626;">Total egresos: <strong>{cls._format_amount(data['total_egresos'])} XAF</strong></div>
        <div style="color:#1e40af;font-size:13px;">Diferencia: <strong>{cls._format_amount(data['diferencia'])} XAF</strong></div>
    </div>

    <div class="footer">
        Generado automáticamente — {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sistema de Gestión de Flujo de Caja
    </div>
</body></html>"""

        try:
            from weasyprint import HTML as WeasyHTML
            return WeasyHTML(string=html).write_pdf()
        except Exception:
            # Fallback: devolver HTML como bytes si WeasyPrint no disponible
            return html.encode('utf-8')

    @classmethod
    def generate_session_xlsx(cls, db: Session, session_id: int) -> Optional[bytes]:
        """Genera informe de cierre de sesión en Excel."""
        data = cls.get_session_data(db, session_id)
        if not data:
            return None

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"Sesión {session_id}"

        # Estilos
        header_font = Font(bold=True, size=14, color="1E40AF")
        sub_font = Font(bold=True, size=10, color="475569")
        th_font = Font(bold=True, size=9, color="FFFFFF")
        th_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        border = Border(bottom=Side(style='thin', color='E2E8F0'))

        # Cabecera
        s = data["session"]
        ws.merge_cells('A1:G1')
        ws['A1'] = "Informe de Cierre de Sesión"
        ws['A1'].font = header_font

        ws['A3'] = "Delegación:"
        ws['B3'] = s['delegacion']
        ws['A4'] = "Gestor:"
        ws['B4'] = s['gestor']
        ws['D3'] = "Apertura:"
        ws['E3'] = s['opened_at']
        ws['D4'] = "Cierre:"
        ws['E4'] = s['closed_at']
        ws['A5'] = "Saldo apertura:"
        ws['B5'] = s['opening_balance']
        ws['D5'] = "Saldo cierre:"
        ws['E5'] = s['closing_balance'] if s['closing_balance'] is not None else 'N/A'

        for cell in ['A3','A4','A5','D3','D4','D5']:
            ws[cell].font = Font(bold=True, size=9)

        # Encabezados de tabla
        headers = ['Fecha/Hora', 'Referencia', 'Concepto', 'Categoría', 'Contraparte', 'Importe', 'Estado']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=h)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for i, t in enumerate(data["transactions"], 8):
            ws.cell(row=i, column=1, value=t['fecha'])
            ws.cell(row=i, column=2, value=t['referencia'])
            ws.cell(row=i, column=3, value=t['concepto'])
            ws.cell(row=i, column=4, value=t['categoria'])
            ws.cell(row=i, column=5, value=t['contraparte'])
            amt_cell = ws.cell(row=i, column=6, value=t['importe'])
            amt_cell.number_format = '#,##0'
            if t['tipo'] == 'income':
                amt_cell.font = Font(color="16A34A", bold=True)
            else:
                amt_cell.font = Font(color="DC2626", bold=True)
            ws.cell(row=i, column=7, value=t['estado'])
            for col in range(1, 8):
                ws.cell(row=i, column=col).border = border

        # Totales
        last_row = 8 + len(data["transactions"])
        ws.cell(row=last_row + 1, column=5, value="Total ingresos:").font = Font(bold=True)
        ws.cell(row=last_row + 1, column=6, value=data['total_ingresos']).font = Font(color="16A34A", bold=True)
        ws.cell(row=last_row + 2, column=5, value="Total egresos:").font = Font(bold=True)
        ws.cell(row=last_row + 2, column=6, value=data['total_egresos']).font = Font(color="DC2626", bold=True)
        ws.cell(row=last_row + 3, column=5, value="Diferencia:").font = Font(bold=True, size=11)
        ws.cell(row=last_row + 3, column=6, value=data['diferencia']).font = Font(bold=True, size=11, color="1E40AF")

        for col_num in [6]:
            for row in [last_row+1, last_row+2, last_row+3]:
                ws.cell(row=row, column=col_num).number_format = '#,##0'

        # Ajustar anchos
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def generate_period_pdf(
        cls, db: Session,
        delegacion: Optional[str],
        date_start: date,
        date_end: date,
        category_id: Optional[int] = None,
        project_id: Optional[int] = None
    ) -> bytes:
        """Genera informe de período libre en PDF."""
        data = cls.get_period_data(db, delegacion, date_start, date_end, category_id, project_id)
        p = data["periodo"]

        rows_html = ""
        for t in data["transactions"]:
            color = "#16a34a" if t["tipo"] == "income" else "#dc2626"
            if t["cancelled"]:
                color = "#9ca3af"
            signo = "+" if t["tipo"] == "income" else "-"
            rows_html += f"""
            <tr style="{'text-decoration:line-through;color:#9ca3af;' if t['cancelled'] else ''}">
                <td>{t['fecha']}</td>
                <td>{t['referencia']}</td>
                <td>{t['concepto'][:60]}</td>
                <td>{t['categoria']}</td>
                <td>{t['contraparte'][:30]}</td>
                <td style="color:{color};text-align:right;font-weight:600;">{signo} {cls._format_amount(t['importe'])} XAF</td>
                <td>{t['delegacion']}</td>
                <td>{t['estado']}</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    @page {{ size: A4 landscape; margin: 15mm; }}
    body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10px; color: #1e293b; }}
    h1 {{ font-size: 16px; color: #1e40af; margin-bottom: 5px; }}
    h2 {{ font-size: 12px; color: #475569; margin-top: 0; }}
    .info {{ background: #f1f5f9; padding: 8px 12px; border-radius: 4px; margin-bottom: 15px; font-size: 10px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
    th {{ background: #1e40af; color: white; padding: 6px 8px; text-align: left; font-size: 9px; }}
    td {{ padding: 5px 8px; border-bottom: 1px solid #e2e8f0; font-size: 9px; }}
    tr:nth-child(even) {{ background: #f8fafc; }}
    .totals {{ margin-top: 15px; text-align: right; }}
    .totals div {{ margin: 3px 0; font-size: 11px; }}
    .footer {{ margin-top: 20px; font-size: 8px; color: #94a3b8; text-align: center; }}
</style></head><body>
    <h1>Informe de Período</h1>
    <h2>{p['date_start']} — {p['date_end']}</h2>
    <div class="info">
        <strong>Delegación:</strong> {p['delegacion']} &nbsp; | &nbsp;
        <strong>Filtros:</strong> {p['filtros']} &nbsp; | &nbsp;
        <strong>Total transacciones:</strong> {len(data['transactions'])}
    </div>

    <table>
        <thead><tr>
            <th>Fecha/Hora</th><th>Ref.</th><th>Concepto</th><th>Categoría</th>
            <th>Contraparte</th><th style="text-align:right">Importe</th><th>Deleg.</th><th>Estado</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
    </table>

    <div class="totals">
        <div style="color:#16a34a;">Total ingresos: <strong>{cls._format_amount(data['total_ingresos'])} XAF</strong></div>
        <div style="color:#dc2626;">Total egresos: <strong>{cls._format_amount(data['total_egresos'])} XAF</strong></div>
        <div style="color:#1e40af;font-size:13px;">Saldo neto: <strong>{cls._format_amount(data['saldo_neto'])} XAF</strong></div>
    </div>

    <div class="footer">
        Generado automáticamente — {datetime.now().strftime('%d/%m/%Y %H:%M')} — Sistema de Gestión de Flujo de Caja
    </div>
</body></html>"""

        try:
            from weasyprint import HTML as WeasyHTML
            return WeasyHTML(string=html).write_pdf()
        except Exception:
            return html.encode('utf-8')

    @classmethod
    def generate_period_xlsx(
        cls, db: Session,
        delegacion: Optional[str],
        date_start: date,
        date_end: date,
        category_id: Optional[int] = None,
        project_id: Optional[int] = None
    ) -> bytes:
        """Genera informe de período libre en Excel."""
        data = cls.get_period_data(db, delegacion, date_start, date_end, category_id, project_id)

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Informe de Período"

        header_font = Font(bold=True, size=14, color="1E40AF")
        th_font = Font(bold=True, size=9, color="FFFFFF")
        th_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        border = Border(bottom=Side(style='thin', color='E2E8F0'))

        p = data["periodo"]
        ws.merge_cells('A1:H1')
        ws['A1'] = "Informe de Período"
        ws['A1'].font = header_font

        ws['A3'] = "Período:"
        ws['B3'] = f"{p['date_start']} — {p['date_end']}"
        ws['A4'] = "Delegación:"
        ws['B4'] = p['delegacion']
        ws['A5'] = "Filtros:"
        ws['B5'] = p['filtros']
        for cell in ['A3','A4','A5']:
            ws[cell].font = Font(bold=True, size=9)

        headers = ['Fecha/Hora', 'Referencia', 'Concepto', 'Categoría', 'Contraparte', 'Importe', 'Delegación', 'Estado']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=h)
            cell.font = th_font
            cell.fill = th_fill
            cell.alignment = Alignment(horizontal='center')

        for i, t in enumerate(data["transactions"], 8):
            ws.cell(row=i, column=1, value=t['fecha'])
            ws.cell(row=i, column=2, value=t['referencia'])
            ws.cell(row=i, column=3, value=t['concepto'])
            ws.cell(row=i, column=4, value=t['categoria'])
            ws.cell(row=i, column=5, value=t['contraparte'])
            amt = ws.cell(row=i, column=6, value=t['importe'])
            amt.number_format = '#,##0'
            if t['tipo'] == 'income':
                amt.font = Font(color="16A34A", bold=True)
            else:
                amt.font = Font(color="DC2626", bold=True)
            ws.cell(row=i, column=7, value=t['delegacion'])
            ws.cell(row=i, column=8, value=t['estado'])
            for col in range(1, 9):
                ws.cell(row=i, column=col).border = border

        last_row = 8 + len(data["transactions"])
        ws.cell(row=last_row+1, column=5, value="Total ingresos:").font = Font(bold=True)
        ws.cell(row=last_row+1, column=6, value=data['total_ingresos']).font = Font(color="16A34A", bold=True)
        ws.cell(row=last_row+2, column=5, value="Total egresos:").font = Font(bold=True)
        ws.cell(row=last_row+2, column=6, value=data['total_egresos']).font = Font(color="DC2626", bold=True)
        ws.cell(row=last_row+3, column=5, value="Saldo neto:").font = Font(bold=True, size=11)
        ws.cell(row=last_row+3, column=6, value=data['saldo_neto']).font = Font(bold=True, size=11, color="1E40AF")

        for row in [last_row+1, last_row+2, last_row+3]:
            ws.cell(row=row, column=6).number_format = '#,##0'

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ============================================================
    # M11 - Recibo individual de transaccion (F-extension)
    # ============================================================

    @staticmethod
    def _number_to_words_es(n: int) -> str:
        """Convierte un entero positivo a palabras en español.
        Soporta hasta cientos de millones (suficiente para XAF)."""
        if n == 0:
            return "cero"
        UNIDADES = ['', 'uno', 'dos', 'tres', 'cuatro', 'cinco', 'seis', 'siete',
                    'ocho', 'nueve', 'diez', 'once', 'doce', 'trece', 'catorce',
                    'quince', 'dieciseis', 'diecisiete', 'dieciocho', 'diecinueve',
                    'veinte', 'veintiuno', 'veintidos', 'veintitres', 'veinticuatro',
                    'veinticinco', 'veintiseis', 'veintisiete', 'veintiocho', 'veintinueve']
        DECENAS = ['', '', '', 'treinta', 'cuarenta', 'cincuenta', 'sesenta',
                   'setenta', 'ochenta', 'noventa']
        CENTENAS = ['', 'ciento', 'doscientos', 'trescientos', 'cuatrocientos',
                    'quinientos', 'seiscientos', 'setecientos', 'ochocientos', 'novecientos']

        def menor_1000(n):
            if n < 30:
                return UNIDADES[n]
            if n < 100:
                d, u = divmod(n, 10)
                if u == 0:
                    return DECENAS[d]
                return DECENAS[d] + ' y ' + UNIDADES[u]
            if n == 100:
                return 'cien'
            cc, r = divmod(n, 100)
            if r == 0:
                return CENTENAS[cc]
            return CENTENAS[cc] + ' ' + menor_1000(r)

        def convertir(n):
            if n < 1000:
                return menor_1000(n)
            if n < 1_000_000:
                miles, resto = divmod(n, 1000)
                pre = 'mil' if miles == 1 else menor_1000(miles) + ' mil'
                return pre if resto == 0 else pre + ' ' + menor_1000(resto)
            if n < 1_000_000_000:
                millones, resto = divmod(n, 1_000_000)
                pre = 'un millon' if millones == 1 else convertir(millones) + ' millones'
                return pre if resto == 0 else pre + ' ' + convertir(resto)
            return str(n)

        return convertir(n)

    @classmethod
    def generate_transaction_receipt_pdf(cls, db: Session, txn_id: int,
                                          requesting_user: User):
        """Genera un recibo PDF de una transaccion individual.
        Retorna bytes (PDF) o None si no existe / sin acceso."""
        txn = db.query(Transaction).filter(Transaction.id == txn_id).first()
        if not txn:
            return None
        # Control de acceso: gestor solo su delegacion
        if requesting_user.role == 'gestor' and txn.delegacion != requesting_user.delegacion:
            return None

        # Cargar relaciones
        cat = db.query(TransactionCategory).filter(TransactionCategory.id == txn.category_id).first()
        subcat = db.query(TransactionSubcategory).filter(TransactionSubcategory.id == txn.subcategory_id).first()
        gestor = db.query(User).filter(User.id == txn.user_id).first()
        contraparte = cls._get_contraparte(db, txn) or '(sin contraparte)'

        # Proyectos asociados
        tps = db.query(TransactionProject).filter(TransactionProject.transaction_id == txn.id).all()
        proyectos_html = ""
        for tp in tps:
            proj = db.query(Project).filter(Project.id == tp.project_id).first()
            work = db.query(Work).filter(Work.id == tp.work_id).first()
            proyectos_html += f"{proj.name if proj else '?'} / {work.name if work else '?'}<br>"
        if not proyectos_html:
            proyectos_html = "(sin proyecto)"

        # Vehiculo opcional
        vehiculo_html = ""
        if txn.vehicle_id:
            v = db.query(Vehicle).filter(Vehicle.id == txn.vehicle_id).first()
            if v:
                vehiculo_html = f"{v.plate} {v.brand or ''}".strip()

        # Importe en letras
        amount_int = int(txn.amount)
        amount_words = cls._number_to_words_es(amount_int).capitalize()
        tipo_label = "INGRESO" if txn.type == "income" else "EGRESO"
        signo = "+" if txn.type == "income" else "-"
        color_tipo = "#16a34a" if txn.type == "income" else "#dc2626"

        # Firma
        signature = db.query(TransactionSignature).filter(
            TransactionSignature.transaction_id == txn.id
        ).first()

        sig_provisional_banner = ""
        if signature:
            metodo = signature.signature_method or "wacom"
            firmante = signature.signer_name or "(sin nombre)"
            firmado_at = signature.signed_at.strftime("%d/%m/%Y %H:%M") if signature.signed_at else ""
            hash_corto = (signature.sha256_hash or "")[:16] + "..." if signature.sha256_hash else "(no hash)"

            if metodo == "wacom_provisional":
                sig_provisional_banner = """
        <div style="background:#fee2e2; border:2px solid #dc2626; padding:8px;
                    text-align:center; margin:10px 0; font-weight:bold; color:#991b1b;">
            FIRMA PROVISIONAL - PENDIENTE DE VERIFICACION BIOMETRICA
        </div>"""

            if metodo == "fingerprint":
                # Sello biometrico
                dedo = signature.fingerprint_finger_position or "(dedo no registrado)"
                score = signature.fingerprint_score or 0
                sig_body = f"""
                <div style="border:2px solid #1e40af; padding:15px; text-align:center;
                            background:#eff6ff; border-radius:6px;">
                    <div style="font-size:24px; color:#1e40af;">[X] VERIFICADO BIOMETRICAMENTE</div>
                    <div style="margin-top:8px; font-size:12px;">
                        Dedo: <strong>{dedo}</strong> &nbsp;|&nbsp;
                        Calidad de match: <strong>{score}/100</strong>
                    </div>
                </div>"""
            else:
                # wacom o wacom_provisional: imagen de la firma
                if signature.signature_data:
                    sig_body = f"""
                <div style="border:1px solid #cbd5e1; padding:10px; text-align:center; background:#fff;">
                    <img src="data:image/png;base64,{signature.signature_data}"
                         style="max-height:120px; max-width:100%;" />
                </div>"""
                else:
                    sig_body = "<p>(firma no disponible)</p>"

            sig_html = f"""
            {sig_provisional_banner}
            <h3 style="margin-top:15px; color:#1e40af; font-size:13px;">Firma del firmante</h3>
            {sig_body}
            <table style="width:100%; margin-top:8px; font-size:10px;">
                <tr>
                    <td><strong>Firmante:</strong> {firmante}</td>
                    <td style="text-align:right;"><strong>Fecha firma:</strong> {firmado_at}</td>
                </tr>
                <tr>
                    <td colspan="2" style="font-size:9px; color:#64748b;">
                        Metodo: {metodo} &nbsp;|&nbsp; SHA-256: {hash_corto}
                    </td>
                </tr>
            </table>"""
        else:
            sig_html = """<div style="background:#fef3c7; border:1px solid #f59e0b;
                          padding:10px; text-align:center; color:#92400e;">
                          Esta transaccion no tiene firma registrada.
                          </div>"""

        # HTML completo
        fecha_txn = txn.created_at.strftime("%d/%m/%Y %H:%M") if txn.created_at else ""
        fecha_recibo = datetime.utcnow().strftime("%d/%m/%Y %H:%M")
        gestor_nombre = gestor.full_name if gestor else "?"

        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    @page {{ size: A4 portrait; margin: 15mm; }}
    body {{ font-family: Arial, Helvetica, sans-serif; font-size: 11px; color: #1e293b; }}
    .header {{ text-align:center; border-bottom:3px solid #1e40af; padding-bottom:10px; margin-bottom:15px; }}
    .header h1 {{ font-size:20px; color:#1e40af; margin:0; }}
    .header h2 {{ font-size:13px; color:#475569; margin:5px 0 0 0; font-weight:normal; }}
    .header .delegacion {{ font-size:11px; color:#64748b; margin-top:3px; }}
    .ref-box {{ background:#1e40af; color:#fff; padding:8px 12px; text-align:center;
                font-size:13px; font-weight:bold; margin-bottom:15px; }}
    table.datos {{ width:100%; border-collapse:collapse; margin-bottom:10px; }}
    table.datos td {{ padding:6px 8px; border-bottom:1px solid #e2e8f0; vertical-align:top; }}
    table.datos td.label {{ width:35%; font-weight:600; color:#475569; background:#f8fafc; }}
    .importe-grande {{ font-size:24px; font-weight:bold; text-align:center;
                       padding:12px; background:#f1f5f9; border-radius:6px;
                       color:{color_tipo}; margin:10px 0; }}
    .importe-letras {{ font-style:italic; text-align:center; padding:6px;
                       font-size:11px; color:#475569; }}
    .footer {{ margin-top:20px; padding-top:10px; border-top:1px solid #cbd5e1;
               font-size:9px; color:#64748b; text-align:center; }}
    .legal {{ margin-top:8px; font-size:8px; color:#94a3b8; }}
</style>
</head><body>
    <div class="header">
        <h1>R2i Network</h1>
        <h2>Recibo de Transaccion</h2>
        <div class="delegacion">Delegacion: {txn.delegacion}</div>
    </div>

    <div class="ref-box">
        Ref.: {txn.reference_number} &nbsp;|&nbsp; {tipo_label}
    </div>

    <div class="importe-grande">{signo} {amount_int:,} XAF</div>
    <div class="importe-letras">{amount_words} francos CFA</div>

    <table class="datos">
        <tr><td class="label">Fecha de la transaccion</td><td>{fecha_txn}</td></tr>
        <tr><td class="label">Concepto</td><td>{txn.concept}</td></tr>
        <tr><td class="label">Categoria / Subcategoria</td>
            <td>{cat.name if cat else '?'} / {subcat.name if subcat else '?'}</td></tr>
        <tr><td class="label">Proyecto / Obra</td><td>{proyectos_html}</td></tr>
        {'<tr><td class="label">Vehiculo</td><td>' + vehiculo_html + '</td></tr>' if vehiculo_html else ''}
        <tr><td class="label">Contraparte</td><td>{contraparte}</td></tr>
    </table>

    {sig_html}

    <div class="footer">
        Registrado por: <strong>{gestor_nombre}</strong>
        &nbsp;|&nbsp; Recibo generado: {fecha_recibo}
        <div class="legal">
            Este documento es generado automaticamente por el sistema Caja R2i.
            La firma incluida es electronica y la integridad del registro puede
            verificarse mediante el hash SHA-256 del documento original.
        </div>
    </div>
</body></html>"""

        try:
            from weasyprint import HTML as WeasyHTML
            return WeasyHTML(string=html).write_pdf()
        except Exception as e:
            # Fallback HTML si WeasyPrint falla
            return html.encode('utf-8')

