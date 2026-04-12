"""
Servicio de importación desde Excel.
Gestiona la validación completa del fichero, la creación de sesión de importación,
la inserción de transacciones con flags de importado y el cálculo de hashes SHA-256.
"""
import io
import re
from datetime import datetime, timedelta
from typing import Optional
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import func as sa_func

from app.config import settings
from app.models.catalogs import (
    Project, Work, TransactionCategory, TransactionSubcategory
)
from app.models.cash_flow import (
    CashSession, Transaction, TransactionProject
)
from app.models.import_history import ImportHistory
from app.models.audit_log import AuditLog
from app.services.integrity_service import compute_transaction_hash


# Columnas esperadas en el Excel (orden flexible, se buscan por nombre)
EXPECTED_HEADERS = [
    "Fecha", "Ref", "Concepto", "Categoría", "Subcategoría",
    "Tipo", "Contraparte", "Monto", "Proyecto", "Obra"
]

# Patrón de referencia: B0001-B9999 o M0001-M9999
REF_PATTERN = re.compile(r'^[BM]\d{4}$')


class ImportService:
    """Servicio de importación de datos históricos desde ficheros Excel."""

    @staticmethod
    def _read_excel(file_bytes: bytes) -> list:
        """Lee el Excel y devuelve lista de dicts con los datos de cada fila."""
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=False)

        # Leer cabecera
        header_row = next(rows_iter, None)
        if not header_row:
            return []

        headers = [cell.value.strip() if cell.value else "" for cell in header_row]

        # Mapear índices de columnas por nombre
        col_map = {}
        for expected in EXPECTED_HEADERS:
            for idx, h in enumerate(headers):
                if h.lower() == expected.lower():
                    col_map[expected] = idx
                    break

        data = []
        for row_num, row in enumerate(rows_iter, start=2):
            values = [cell.value for cell in row]
            # Saltar filas completamente vacías
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
                continue
            row_dict = {"_row_num": row_num}
            for col_name, idx in col_map.items():
                row_dict[col_name] = values[idx] if idx < len(values) else None
            data.append(row_dict)

        wb.close()
        return data

    @staticmethod
    def _validate_row(row: dict, db: Session, delegacion: str,
                      existing_refs: set, categories_cache: dict,
                      subcategories_cache: dict) -> list:
        """Valida una fila del Excel. Devuelve lista de errores (vacía si válida)."""
        errors = []
        row_num = row["_row_num"]

        # Fecha
        fecha = row.get("Fecha")
        if fecha is None:
            errors.append({"row": row_num, "field": "Fecha", "error": "Fecha vacía", "value": None})
        else:
            if isinstance(fecha, datetime):
                fecha_dt = fecha
            elif isinstance(fecha, str):
                try:
                    fecha_dt = datetime.strptime(fecha.strip(), "%Y-%m-%d")
                except ValueError:
                    try:
                        fecha_dt = datetime.strptime(fecha.strip(), "%d/%m/%Y")
                    except ValueError:
                        errors.append({"row": row_num, "field": "Fecha",
                                       "error": "Formato de fecha no válido", "value": str(fecha)})
                        fecha_dt = None
            else:
                errors.append({"row": row_num, "field": "Fecha",
                               "error": "Tipo de dato no válido para fecha", "value": str(fecha)})
                fecha_dt = None

            if fecha_dt and fecha_dt.date() > datetime.now().date():
                errors.append({"row": row_num, "field": "Fecha",
                               "error": "La fecha no puede ser posterior a hoy", "value": str(fecha)})

        # Referencia
        ref = row.get("Ref")
        if ref is None or (isinstance(ref, str) and ref.strip() == ""):
            errors.append({"row": row_num, "field": "Ref", "error": "Referencia vacía", "value": None})
        else:
            ref_str = str(ref).strip().upper()
            if not REF_PATTERN.match(ref_str):
                errors.append({"row": row_num, "field": "Ref",
                               "error": "Formato inválido (esperado B#### o M####)", "value": ref_str})

        # Concepto
        concepto = row.get("Concepto")
        if not concepto or (isinstance(concepto, str) and len(concepto.strip()) < 3):
            errors.append({"row": row_num, "field": "Concepto",
                           "error": "Concepto vacío o menor a 3 caracteres",
                           "value": str(concepto) if concepto else None})

        # Categoría
        categoria_name = row.get("Categoría")
        if not categoria_name or (isinstance(categoria_name, str) and categoria_name.strip() == ""):
            errors.append({"row": row_num, "field": "Categoría", "error": "Categoría vacía", "value": None})
        else:
            cat_name = str(categoria_name).strip()
            if cat_name not in categories_cache:
                errors.append({"row": row_num, "field": "Categoría",
                               "error": f"Categoría '{cat_name}' no existe en el catálogo",
                               "value": cat_name})

        # Subcategoría
        subcategoria_name = row.get("Subcategoría")
        if not subcategoria_name or (isinstance(subcategoria_name, str) and subcategoria_name.strip() == ""):
            errors.append({"row": row_num, "field": "Subcategoría",
                           "error": "Subcategoría vacía", "value": None})
        else:
            sub_name = str(subcategoria_name).strip()
            cat_name = str(categoria_name).strip() if categoria_name else ""
            if cat_name in categories_cache:
                cat_id = categories_cache[cat_name]
                key = (cat_id, sub_name)
                if key not in subcategories_cache:
                    errors.append({"row": row_num, "field": "Subcategoría",
                                   "error": f"Subcategoría '{sub_name}' no existe para categoría '{cat_name}'",
                                   "value": sub_name})

        # Tipo
        tipo = row.get("Tipo")
        if not tipo or str(tipo).strip().lower() not in ("ingreso", "egreso"):
            errors.append({"row": row_num, "field": "Tipo",
                           "error": "Tipo debe ser 'Ingreso' o 'Egreso'",
                           "value": str(tipo) if tipo else None})

        # Contraparte
        contraparte = row.get("Contraparte")
        if not contraparte or (isinstance(contraparte, str) and len(contraparte.strip().split()) < 2):
            errors.append({"row": row_num, "field": "Contraparte",
                           "error": "Contraparte vacía o sin apellido (mínimo dos palabras)",
                           "value": str(contraparte) if contraparte else None})

        # Monto
        monto = row.get("Monto")
        if monto is None:
            errors.append({"row": row_num, "field": "Monto", "error": "Monto vacío", "value": None})
        else:
            try:
                monto_val = float(str(monto).replace(",", ".").replace(" ", ""))
                if monto_val <= 0:
                    errors.append({"row": row_num, "field": "Monto",
                                   "error": "Monto debe ser positivo", "value": str(monto)})
            except (ValueError, TypeError):
                errors.append({"row": row_num, "field": "Monto",
                               "error": "Monto no es un número válido", "value": str(monto)})

        # Proyecto (no genera error, se crea automáticamente)
        proyecto = row.get("Proyecto")
        if not proyecto or (isinstance(proyecto, str) and proyecto.strip() == ""):
            errors.append({"row": row_num, "field": "Proyecto",
                           "error": "Proyecto vacío", "value": None})

        # Obra (no genera error, se crea automáticamente)
        obra = row.get("Obra")
        if not obra or (isinstance(obra, str) and obra.strip() == ""):
            errors.append({"row": row_num, "field": "Obra",
                           "error": "Obra vacía", "value": None})

        return errors

    @staticmethod
    def validate_excel(db: Session, file_bytes: bytes, delegacion: str) -> dict:
        """
        Valida el fichero Excel completo sin importar nada.
        Devuelve un informe con filas válidas, errores, duplicados y entidades a crear.
        """
        data = ImportService._read_excel(file_bytes)
        if not data:
            return {
                "total_rows": 0, "valid_rows": 0, "error_rows": 0,
                "duplicate_rows": 0, "projects_to_create": [],
                "works_to_create": [], "errors": [], "can_import": False
            }

        # Caches para validación
        categories = db.query(TransactionCategory).all()
        categories_cache = {c.name: c.id for c in categories}

        subcategories = db.query(TransactionSubcategory).all()
        subcategories_cache = {(s.category_id, s.name): s.id for s in subcategories}

        existing_refs = set(
            r[0] for r in db.query(Transaction.reference_number).all()
        )

        projects_db = {p.name: p.id for p in db.query(Project).all()}
        works_db = {}
        for w in db.query(Work).all():
            works_db[(w.project_id, w.name)] = w.id

        all_errors = []
        duplicates = 0
        valid_count = 0
        projects_to_create = set()
        works_to_create = set()

        for row in data:
            ref = row.get("Ref")
            ref_str = str(ref).strip().upper() if ref else ""

            # Detectar duplicados
            if ref_str in existing_refs:
                duplicates += 1
                continue

            row_errors = ImportService._validate_row(
                row, db, delegacion, existing_refs, categories_cache, subcategories_cache
            )

            if row_errors:
                all_errors.extend(row_errors)
            else:
                valid_count += 1

            # Detectar proyectos/obras a crear
            proyecto_name = str(row.get("Proyecto", "")).strip()
            obra_name = str(row.get("Obra", "")).strip()
            if proyecto_name and proyecto_name not in projects_db:
                projects_to_create.add(proyecto_name)
            if proyecto_name and obra_name:
                proj_id = projects_db.get(proyecto_name)
                if proj_id and (proj_id, obra_name) not in works_db:
                    works_to_create.add(f"{proyecto_name} → {obra_name}")
                elif not proj_id:
                    works_to_create.add(f"{proyecto_name} → {obra_name}")

        return {
            "total_rows": len(data),
            "valid_rows": valid_count,
            "error_rows": len(set(e["row"] for e in all_errors)),
            "duplicate_rows": duplicates,
            "projects_to_create": sorted(projects_to_create),
            "works_to_create": sorted(works_to_create),
            "errors": all_errors,
            "can_import": len(all_errors) == 0
        }

    @staticmethod
    def execute_import(db: Session, file_bytes: bytes, delegacion: str,
                       user_id: int, filename: str) -> dict:
        """
        Ejecuta la importación tras validación exitosa.
        Crea sesión de importación, transacciones con flags de importado y hash SHA-256.
        """
        # Validar primero
        validation = ImportService.validate_excel(db, file_bytes, delegacion)
        if not validation["can_import"]:
            raise ValueError("El fichero contiene errores. Validar antes de importar.")

        data = ImportService._read_excel(file_bytes)
        import_edit_days = int(getattr(settings, 'IMPORT_EDIT_DAYS', 30))
        now = datetime.utcnow()

        # Caches
        categories = {c.name: c.id for c in db.query(TransactionCategory).all()}
        subcategories_all = db.query(TransactionSubcategory).all()
        subcategories = {(s.category_id, s.name): s.id for s in subcategories_all}
        projects = {p.name: p for p in db.query(Project).all()}
        works_all = db.query(Work).all()
        works = {}
        for w in works_all:
            works[(w.project_id, w.name)] = w
        existing_refs = set(r[0] for r in db.query(Transaction.reference_number).all())

        # Crear sesión de importación (cerrada desde el inicio)
        import_session = CashSession(
            user_id=user_id,
            delegacion=delegacion,
            opened_at=now,
            closed_at=now,
            opening_balance=0,
            closing_balance=0,
            status="closed",
            notes=f"[IMPORT] {filename}"
        )
        db.add(import_session)
        db.flush()

        rows_imported = 0
        rows_skipped = 0
        projects_created = 0
        works_created = 0

        for row in data:
            ref_str = str(row.get("Ref", "")).strip().upper()

            # Omitir duplicados
            if ref_str in existing_refs:
                rows_skipped += 1
                continue

            # Resolver fecha
            fecha = row.get("Fecha")
            if isinstance(fecha, datetime):
                created_at = fecha
            elif isinstance(fecha, str):
                try:
                    created_at = datetime.strptime(fecha.strip(), "%Y-%m-%d")
                except ValueError:
                    created_at = datetime.strptime(fecha.strip(), "%d/%m/%Y")
            else:
                created_at = now

            # Resolver categoría y subcategoría
            cat_name = str(row.get("Categoría", "")).strip()
            sub_name = str(row.get("Subcategoría", "")).strip()
            category_id = categories.get(cat_name)
            subcategory_id = subcategories.get((category_id, sub_name))

            # Resolver tipo
            tipo_raw = str(row.get("Tipo", "")).strip().lower()
            tx_type = "income" if tipo_raw == "ingreso" else "expense"

            # Resolver monto
            monto_str = str(row.get("Monto", "0")).replace(",", ".").replace(" ", "")
            amount = round(float(monto_str), 2)

            # Resolver proyecto (crear si no existe)
            proyecto_name = str(row.get("Proyecto", "")).strip()
            proyecto_code = proyecto_name.replace(" ", "_")[:30]
            project_obj = projects.get(proyecto_name)
            if project_obj is None:
                # Buscar por code en BD
                project_obj = db.query(Project).filter(
                    (Project.name == proyecto_name) | (Project.code == proyecto_code)
                ).first()
            if project_obj is None:
                project_obj = Project(code=proyecto_code, name=proyecto_name, active=True)
                db.add(project_obj)
                db.flush()
                projects_created += 1
            projects[proyecto_name] = project_obj

            # Resolver obra (crear si no existe)
            obra_name = str(row.get("Obra", "")).strip()
            obra_code = obra_name.replace(" ", "_")[:50]
            work_key = (project_obj.id, obra_name)
            work_key_code = (project_obj.id, obra_code)
            existing_work = works.get(work_key) or works.get(work_key_code)
            if existing_work is None:
                # Buscar en BD por code como última verificación
                existing_work = db.query(Work).filter(
                    Work.project_id == project_obj.id,
                    Work.code == obra_code
                ).first()
            if existing_work is None:
                new_work = Work(
                    project_id=project_obj.id, code=obra_code,
                    name=obra_name, active=True
                )
                db.add(new_work)
                db.flush()
                works[work_key] = new_work
                works[work_key_code] = new_work
                works_created += 1
                existing_work = new_work
            else:
                works[work_key] = existing_work
                works[work_key_code] = existing_work

            work_obj = works[work_key]

            # Contraparte
            contraparte = str(row.get("Contraparte", "")).strip()
            concepto = str(row.get("Concepto", "")).strip()

            # Crear transacción importada
            tx = Transaction(
                session_id=import_session.id,
                delegacion=delegacion,
                category_id=category_id,
                subcategory_id=subcategory_id,
                user_id=user_id,
                counterparty_free=contraparte,
                type=tx_type,
                amount=amount,
                concept=concepto,
                reference_number=ref_str,
                transaction_type="normal",
                imported=True,
                import_source=filename,
                imported_editable_until=now + timedelta(days=import_edit_days),
                editable_until=now,  # Ya expirada para nativos
                approval_status="approved",
                cancelled=False,
                is_adjustment=False,
                integrity_hash="pending",  # Se recalcula después del flush
                created_at=created_at
            )
            db.add(tx)
            db.flush()

            # Calcular hash SHA-256 real
            tx.integrity_hash = compute_transaction_hash(tx)

            # Crear vínculo transacción-proyecto-obra
            tp = TransactionProject(
                transaction_id=tx.id,
                project_id=project_obj.id,
                work_id=work_obj.id
            )
            db.add(tp)

            existing_refs.add(ref_str)
            rows_imported += 1

        # Registrar en historial
        history = ImportHistory(
            delegacion=delegacion,
            filename=filename,
            session_id=import_session.id,
            rows_imported=rows_imported,
            rows_skipped=rows_skipped,
            projects_created=projects_created,
            works_created=works_created,
            imported_by=user_id
        )
        db.add(history)

        # Audit log
        audit = AuditLog(
            user_id=user_id,
            delegacion=delegacion,
            action="IMPORT_EXCEL",
            entity="import_history",
            entity_id=None,
            details={
                "filename": filename,
                "rows_imported": rows_imported,
                "rows_skipped": rows_skipped,
                "projects_created": projects_created,
                "works_created": works_created
            }
        )
        db.add(audit)

        db.commit()

        # Obtener el ID del historial tras commit
        db.refresh(history)

        return {
            "rows_imported": rows_imported,
            "rows_skipped": rows_skipped,
            "projects_created": projects_created,
            "works_created": works_created,
            "session_id": import_session.id,
            "import_id": history.id
        }

    @staticmethod
    def get_history(db: Session) -> list:
        """Devuelve el historial completo de importaciones ordenado por fecha descendente."""
        return (
            db.query(ImportHistory)
            .order_by(ImportHistory.imported_at.desc())
            .all()
        )
